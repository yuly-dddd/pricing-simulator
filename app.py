"""
KL 가격인상 시뮬레이션 v3
유저가 7대 시나리오 변수를 각각 세분화 옵션으로 설정
"""

from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
from datetime import datetime

app = Flask(__name__)

# ══════════════════════════════════════════
# 7대 시나리오 변수 정의 + 프리셋 옵션
# ══════════════════════════════════════════

SCENARIO_VARIABLES = [
    {
        "id": "price_increase",
        "label": "① 가격인상",
        "icon": "💰",
        "description": "소비자가/물대 인상 수준",
        "unit": "억원",
        "presets": [
            {"id": "off", "label": "미인상", "value": 0, "detail": "가격 동결"},
            {"id": "half", "label": "50% 수준", "value": 4029373459, "detail": "버거 150원, 치킨 500원 인상"},
            {"id": "plan", "label": "현재 계획", "value": 8058746918, "detail": "버거 300원, 빅싸이 1,000원, 맥스 1,500원", "default": True},
            {"id": "high", "label": "강하게 인상", "value": 12088120377, "detail": "버거 500원, 빅싸이 1,500원, 맥스 2,000원"},
        ],
        "custom_range": {"min": 0, "max": 20000000000, "step": 500000000},
        "base_impact": 8058746918,
    },
    {
        "id": "volume_decrease",
        "label": "② 판매량 감소",
        "icon": "📉",
        "description": "가격인상에 따른 TC(고객수) 감소",
        "unit": "억원",
        "presets": [
            {"id": "none", "label": "감소 없음", "value": 0, "detail": "TC 변동 없음"},
            {"id": "mild", "label": "소폭 감소", "value": -721605497, "detail": "TC 3.7%, 3개월 감소"},
            {"id": "plan", "label": "현재 계획", "value": -1443210994, "detail": "TC 7.4%, 3개월 감소", "default": True},
            {"id": "severe", "label": "대폭 감소", "value": -2886421988, "detail": "TC 15%, 3개월 감소"},
        ],
        "sub_params": [
            {"id": "tc_rate", "label": "TC 감소율(%)", "default": 7.43, "min": 0, "max": 30, "step": 0.5},
            {"id": "tc_months", "label": "감소기간(개월)", "default": 3, "min": 1, "max": 12, "step": 1},
        ],
        "custom_range": {"min": -5000000000, "max": 0, "step": 100000000},
        "base_impact": -1443210994,
    },
    {
        "id": "store_increase",
        "label": "③ 매장순증",
        "icon": "🏪",
        "description": "신규 매장 오픈에 따른 매출 증분",
        "unit": "억원",
        "presets": [
            {"id": "none", "label": "순증 없음", "value": 0, "detail": "매장수 유지"},
            {"id": "low", "label": "30개 순증", "value": 4501279181, "detail": "터치 25개 + 피자 5개"},
            {"id": "plan", "label": "현재 계획", "value": 7502131968, "detail": "터치 약 50개 순증 (1,453→1,503)", "default": True},
            {"id": "high", "label": "80개 순증", "value": 12003411149, "detail": "터치 70개 + 피자 10개"},
        ],
        "sub_params": [
            {"id": "new_stores", "label": "순증 매장수", "default": 50.5, "min": 0, "max": 150, "step": 5},
        ],
        "custom_range": {"min": 0, "max": 20000000000, "step": 500000000},
        "base_impact": 7502131968,
    },
    {
        "id": "overseas",
        "label": "④ 해외순증",
        "icon": "🌏",
        "description": "해외매장 확대에 따른 로열티/수출 증분",
        "unit": "억원",
        "presets": [
            {"id": "none", "label": "증분 없음", "value": 0, "detail": "해외 현상 유지"},
            {"id": "plan", "label": "현재 계획", "value": 498445256, "detail": "26년 1월 매출 연환산 기준", "default": True},
            {"id": "high", "label": "적극 확대", "value": 996890512, "detail": "MF 및 해외매장 적극 확대"},
        ],
        "custom_range": {"min": 0, "max": 3000000000, "step": 100000000},
        "base_impact": 498445256,
    },
    {
        "id": "beef_expansion",
        "label": "⑤ 비프매출 증대",
        "icon": "🥩",
        "description": "비프버거 MIX 증대 & 설치매장 확대",
        "unit": "억원",
        "presets": [
            {"id": "none", "label": "미반영", "value": 0, "detail": "비프 확대 없음"},
            {"id": "conservative", "label": "보수적", "value": 1440402856, "detail": "MIX 3.7%→4.3%, 매장 200개 확대"},
            {"id": "plan", "label": "현재 계획", "value": 2880805712, "detail": "MIX 3.7%→5.0%, 매장 452개 확대", "default": True},
            {"id": "aggressive", "label": "적극적", "value": 4321208568, "detail": "MIX 3.7%→6.0%, 매장 600개 확대"},
        ],
        "sub_params": [
            {"id": "beef_mix_target", "label": "비프 MIX 목표(%)", "default": 5.0, "min": 3.7, "max": 10, "step": 0.5},
            {"id": "beef_stores", "label": "그리들 설치매장수", "default": 452, "min": 0, "max": 1000, "step": 50},
        ],
        "custom_range": {"min": 0, "max": 8000000000, "step": 200000000},
        "base_impact": 2880805712,
    },
    {
        "id": "marketing_efficiency",
        "label": "⑥ 마케팅 효율",
        "icon": "📣",
        "description": "매출 대비 마케팅비율 개선",
        "unit": "억원",
        "presets": [
            {"id": "none", "label": "개선 없음", "value": 0, "detail": "마케팅비율 유지 (3.16%)"},
            {"id": "plan", "label": "현재 계획", "value": 311569324, "detail": "0.3%p 개선 (3.16%→2.86%)", "default": True},
            {"id": "target", "label": "Target 달성", "value": 2895336075, "detail": "과거 6Q 평균 Target (2.39%) 달성"},
        ],
        "sub_params": [
            {"id": "mkt_improvement", "label": "마케팅비 개선(%p)", "default": 0.3, "min": 0, "max": 1.5, "step": 0.1},
        ],
        "custom_range": {"min": 0, "max": 5000000000, "step": 100000000},
        "base_impact": 311569324,
    },
    {
        "id": "cost_reduction",
        "label": "⑦ 매입가인하",
        "icon": "📦",
        "description": "원재료(닭고기) 매입단가 인하 전망",
        "unit": "억원",
        "presets": [
            {"id": "none", "label": "인하 없음", "value": 0, "detail": "현재 매입가 유지"},
            {"id": "mild", "label": "소폭 인하", "value": 2487225308, "detail": "오퍼가 7% 인하 (10월부터)"},
            {"id": "plan", "label": "현재 계획", "value": 4974450615, "detail": "오퍼가 2.2불 수준 (13% 인하, 8월부터)", "default": True},
            {"id": "full", "label": "대폭 인하", "value": 7461675923, "detail": "오퍼가 2.0불 수준 (20% 인하, 7월부터)"},
        ],
        "sub_params": [
            {"id": "cost_reduction_pct", "label": "매입가 인하율(%)", "default": 13, "min": 0, "max": 30, "step": 1},
            {"id": "cost_reduction_month", "label": "적용 시작월", "default": 8, "min": 1, "max": 12, "step": 1},
        ],
        "custom_range": {"min": 0, "max": 12000000000, "step": 500000000},
        "base_impact": 4974450615,
    },
]

# ── 글로벌 파라미터 ──
GLOBAL_PARAMS = [
    {"id": "increase_month", "label": "가격인상 시점(월)", "default": 3, "min": 1, "max": 12},
    {"id": "hq_share", "label": "본사 분담비중(%)", "default": 50, "min": 0, "max": 100},
    {"id": "cpi_rate", "label": "물가상승률(%)", "default": 1.31, "min": 0, "max": 10, "step": 0.1},
    {"id": "wage_rate", "label": "임금상승률(%)", "default": 4.0, "min": 0, "max": 15, "step": 0.5},
    {"id": "base_cost_ratio", "label": "Base 원가율(%)", "default": 61.83, "min": 50, "max": 80, "step": 0.5},
]

# ── 손익계산서 ──
PNL = {
    "revenue":              {"label": "총매출액",         "y25": 479876076315},
    "revenue_franchise":    {"label": "  가맹매출액",     "y25": 415129789902},
    "revenue_franchise_t":  {"label": "    맘스터치",     "y25": 406465489141},
    "revenue_franchise_p":  {"label": "    맘스피자",     "y25": 8664300761},
    "revenue_direct":       {"label": "  직영매출액",     "y25": 20888561841},
    "revenue_dist":         {"label": "  유통사업",       "y25": 41457001701},
    "revenue_other":        {"label": "  기타매출액",     "y25": 2400722871},
    "cogs":                 {"label": "매출원가",         "y25": 300097805571},
    "gross_profit":         {"label": "매출총이익",       "y25": 179778270744, "bold": True},
    "variable_cost":        {"label": "변동비",           "y25": 42241846335},
    "marketing":            {"label": "  마케팅비용",     "y25": 15186857123},
    "commission":           {"label": "  지사수수료",     "y25": 12440181710},
    "delivery":             {"label": "  운반비",         "y25": 14614807502},
    "contribution":         {"label": "공헌이익",         "y25": 137536424409, "bold": True},
    "fixed_cost":           {"label": "고정비용",         "y25": 50828153517},
    "op_profit":            {"label": "영업이익",         "y25": 86708270892, "bold": True},
    "subsidiary":           {"label": "자회사손익",       "y25": 2979288745},
    "tokyo":                {"label": "도쿄법인손익",     "y25": -4188390212},
    "consol_op":            {"label": "연결 영업이익",    "y25": 89687559637, "bold": True},
    "da":                   {"label": "D&A",              "y25": 11896630557},
    "ebitda":               {"label": "연결 EBITDA",      "y25": 101584190194, "bold": True, "hl": True},
}

PLAN_Y26 = {
    "revenue": 519253356306, "revenue_franchise": 453836766528,
    "revenue_franchise_t": 440016171240, "revenue_franchise_p": 13820595288,
    "revenue_direct": 20888561841, "revenue_dist": 41457001701, "revenue_other": 3071026236,
    "cogs": 317003716087, "gross_profit": 202249640219,
    "variable_cost": 44150333267, "marketing": 14875287799,
    "commission": 13460988003, "delivery": 15814057466,
    "contribution": 158099306952, "fixed_cost": 51508379856,
    "op_profit": 106590927096, "subsidiary": 3731303228, "tokyo": -4188390212,
    "consol_op": 110322230323, "da": 11896630557, "ebitda": 122218860880,
}

GP_PCT = {"y25": 37.46, "y26_plan": 38.95}
GP_FRANCHISE = {"y25": 37.58, "y26_plan": 39.27}
EBITDA_PCT = {"y25": 21.17, "y26_plan": 23.54}
OP_MARGIN = {"y25": 18.07, "y26_plan": 20.53}

# ── 경쟁사 데이터 (동일) ──
COMPETITORS = {
    "burger_single": [
        {"brand":"맘스터치","product":"싸이버거","weight":256,"current":4900,"after":5200,"ppg_cur":19.1,"ppg_aft":20.3,"ours":True},
        {"brand":"롯데리아","product":"불고기버거","weight":203,"current":5000,"after":5300,"ppg_cur":24.6,"ppg_aft":26.1},
        {"brand":"맥도날드","product":"빅맥","weight":224,"current":5500,"after":5800,"ppg_cur":24.6,"ppg_aft":25.9},
        {"brand":"KFC","product":"징거버거","weight":222,"current":5900,"after":6200,"ppg_cur":26.6,"ppg_aft":27.9},
        {"brand":"버거킹","product":"와퍼","weight":305,"current":7200,"after":7500,"ppg_cur":23.6,"ppg_aft":24.6},
    ],
    "chicken_boneless": [
        {"brand":"맘스터치","product":"빅싸이순살맥스","weight":760,"current":19900,"after":21400,"ppg_cur":26.2,"ppg_aft":28.2,"ours":True},
        {"brand":"맘스터치","product":"빅싸이순살","weight":380,"current":11900,"after":12900,"ppg_cur":31.3,"ppg_aft":33.9,"ours":True},
        {"brand":"BHC","product":"콰삭킹 순살","weight":476,"current":23000,"after":24500,"ppg_cur":48.3,"ppg_aft":51.5},
        {"brand":"BBQ","product":"황금올리브 순살","weight":705,"current":25000,"after":26500,"ppg_cur":35.5,"ppg_aft":37.6},
        {"brand":"교촌","product":"후라이드 순살","weight":457,"current":22000,"after":23500,"ppg_cur":48.1,"ppg_aft":51.4},
    ],
}

QUARTERLY_GP = [
    {"p":"24.1Q","v":35.82},{"p":"24.2Q","v":35.97},{"p":"24.3Q","v":35.33},{"p":"24.4Q","v":38.09},
    {"p":"25.1Q","v":38.68},{"p":"25.2Q","v":38.05},{"p":"25.3Q","v":36.84},{"p":"25.4Q","v":36.73},
]


def simulate(var_values, global_params):
    """7개 변수값을 받아 EBITDA 계산"""
    e25 = PNL["ebitda"]["y25"]
    total_impact = sum(var_values.values())
    e26 = e25 + total_impact

    # Revenue/cost 추정
    rev_25 = PNL["revenue"]["y25"]
    fr_25 = PNL["revenue_franchise"]["y25"]
    cogs_25 = PNL["cogs"]["y25"]

    # 가격인상 → 매출 증가
    price_rev = var_values.get("price_increase", 0) / 0.55 if var_values.get("price_increase", 0) > 0 else 0
    # 매장순증 → 매출 증가
    store_rev = var_values.get("store_increase", 0) / 0.38 if var_values.get("store_increase", 0) > 0 else 0
    # 비프 → 매출 증가
    beef_rev = var_values.get("beef_expansion", 0) / 0.36 if var_values.get("beef_expansion", 0) > 0 else 0
    # 해외 → 매출 증가
    overseas_rev = var_values.get("overseas", 0) / 0.74 if var_values.get("overseas", 0) > 0 else 0
    # 판매량감소 → 매출 감소
    vol_rev = var_values.get("volume_decrease", 0) / 0.38 if var_values.get("volume_decrease", 0) < 0 else 0

    rev_change = price_rev + store_rev + beef_rev + overseas_rev + vol_rev
    rev_26 = rev_25 + rev_change + (PLAN_Y26["revenue_other"] - PNL["revenue_other"]["y25"])

    cogs_change = rev_change * 0.62 - var_values.get("cost_reduction", 0)
    cogs_26 = cogs_25 + cogs_change

    gp_26 = rev_26 - cogs_26
    gp_pct = gp_26 / rev_26 * 100 if rev_26 > 0 else 0

    fr_26 = fr_25 + price_rev + store_rev * 0.7 + beef_rev + vol_rev
    fr_cogs = fr_26 * 0.62 - var_values.get("cost_reduction", 0) * 0.8
    fr_gp_pct = (fr_26 - fr_cogs) / fr_26 * 100 if fr_26 > 0 else 0

    mkt_saving = var_values.get("marketing_efficiency", 0)
    var_cost = PNL["variable_cost"]["y25"] + rev_change * 0.088 - mkt_saving
    contribution = gp_26 - var_cost
    fixed = PNL["fixed_cost"]["y25"] * (1 + global_params.get("wage_rate", 4) / 100 * 0.4)
    op = contribution - fixed
    op_pct = op / rev_26 * 100 if rev_26 > 0 else 0
    ebitda_pct = e26 / rev_26 * 100 if rev_26 > 0 else 0

    return {
        "ebitda_25": e25, "ebitda_26": e26, "ebitda_change": total_impact,
        "ebitda_change_pct": total_impact / e25 * 100,
        "ebitda_pct_25": e25 / rev_25 * 100, "ebitda_pct_26": ebitda_pct,
        "revenue_25": rev_25, "revenue_26": rev_26,
        "gp_pct_25": GP_PCT["y25"], "gp_pct_26": round(gp_pct, 2),
        "franchise_gp_25": GP_FRANCHISE["y25"], "franchise_gp_26": round(fr_gp_pct, 2),
        "op_margin_25": OP_MARGIN["y25"], "op_margin_26": round(op_pct, 2),
        "var_values": var_values,
    }


def build_excel(results_list):
    wb = openpyxl.Workbook()
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    pos = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    neg = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    hl = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
    bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    ws = wb.active; ws.title = "시나리오 비교"
    ws.cell(1,1,"KL 가격인상 시뮬레이션 - 시나리오 비교").font = Font(bold=True, size=14)
    headers = ["시나리오명","①가격인상","②판매량감소","③매장순증","④해외순증","⑤비프확대","⑥마케팅효율","⑦매입가인하",
               "EBITDA합계(억)","증감(억)","EBITDA%","GP%","가맹GP%","OPM%"]
    for c,h in enumerate(headers,1):
        cell = ws.cell(3,c,h); cell.font=hf; cell.fill=hfill; cell.border=bd
    for i,r in enumerate(results_list, 4):
        vv = r["result"]["var_values"]
        data = [r["name"],
                round(vv.get("price_increase",0)/1e8,1), round(vv.get("volume_decrease",0)/1e8,1),
                round(vv.get("store_increase",0)/1e8,1), round(vv.get("overseas",0)/1e8,1),
                round(vv.get("beef_expansion",0)/1e8,1), round(vv.get("marketing_efficiency",0)/1e8,1),
                round(vv.get("cost_reduction",0)/1e8,1),
                round(r["result"]["ebitda_26"]/1e8,0), round(r["result"]["ebitda_change"]/1e8,0),
                round(r["result"]["ebitda_pct_26"],1), round(r["result"]["gp_pct_26"],1),
                round(r["result"]["franchise_gp_26"],1), round(r["result"]["op_margin_26"],1)]
        for c,v in enumerate(data,1):
            cell = ws.cell(i,c,v); cell.border=bd
            if c==10: cell.fill = pos if v>=0 else neg
    for c in range(1,15): ws.column_dimensions[get_column_letter(c)].width=15

    ws2 = wb.create_sheet("P&L 기준"); r=1
    ws2.cell(r,1,"요약 손익계산서 (단위: 백만원)").font=Font(bold=True,size=14); r+=2
    for c,h in enumerate(["구분","2025년","2026년(계획)","증감"],1):
        cell=ws2.cell(r,c,h); cell.font=hf; cell.fill=hfill; cell.border=bd
    r+=1
    for k,item in PNL.items():
        ws2.cell(r,1,item["label"]).border=bd
        ws2.cell(r,2,round(item["y25"]/1e6)).number_format='#,##0'; ws2.cell(r,2).border=bd
        v26=PLAN_Y26.get(k,item["y25"])
        ws2.cell(r,3,round(v26/1e6)).number_format='#,##0'; ws2.cell(r,3).border=bd
        d=round((v26-item["y25"])/1e6)
        ws2.cell(r,4,d).number_format='#,##0'; ws2.cell(r,4).border=bd
        ws2.cell(r,4).fill = pos if d>=0 else neg
        if item.get("bold"): ws2.cell(r,1).font=Font(bold=True)
        if item.get("hl"):
            for c in range(1,5): ws2.cell(r,c).fill=hl
        r+=1
    for c in range(1,5): ws2.column_dimensions[get_column_letter(c)].width=18
    return wb


@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/config")
def get_config():
    return jsonify({
        "variables": SCENARIO_VARIABLES,
        "global_params": GLOBAL_PARAMS,
        "pnl": PNL,
        "plan_y26": PLAN_Y26,
        "gp": GP_PCT, "gp_fr": GP_FRANCHISE, "ebitda_pct": EBITDA_PCT, "op_margin": OP_MARGIN,
        "competitors": COMPETITORS,
        "quarterly_gp": QUARTERLY_GP,
    })

@app.route("/api/simulate", methods=["POST"])
def api_simulate():
    data = request.json
    scenarios = data.get("scenarios", [])
    global_params = data.get("global_params", {})
    results = []
    for sc in scenarios:
        var_values = sc.get("var_values", {})
        res = simulate(var_values, global_params)
        results.append({"name": sc.get("name",""), "result": res})
    return jsonify(results)

@app.route("/api/export", methods=["POST"])
def export_excel():
    data = request.json
    scenarios = data.get("scenarios", [])
    global_params = data.get("global_params", {})
    results_list = []
    for sc in scenarios:
        res = simulate(sc.get("var_values",{}), global_params)
        results_list.append({"name": sc.get("name",""), "result": res})
    wb = build_excel(results_list)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name); tmp.close()
    return send_file(tmp.name, as_attachment=True,
                     download_name=f"KL_시뮬레이션_{datetime.now().strftime('%y%m%d_%H%M')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    print("="*50)
    print("  KL 가격인상 시뮬레이션 v3")
    print("  http://127.0.0.1:5001")
    print("="*50)
    app.run(debug=True, port=5001)
